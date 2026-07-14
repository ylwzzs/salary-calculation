"""Excel 导入器。*_from_rows 版本接收含表头的二维列表，便于单测；
load_*_from_xlsx 用 openpyxl 读真实文件后转调。门店名保留原始，由 calculator.clean_store 统一清洗。"""
from datetime import datetime
from decimal import Decimal

from salary_engine.models import Product, Store, SalesLine


def _D(v) -> Decimal:
    return Decimal(0) if v in (None, "") else Decimal(str(v))


def _norm(s):
    """去掉空白字符，用于表头匹配容错（真实表头常有『名  称』带空格）。"""
    return str(s).replace(" ", "").replace("　", "").replace("\t", "")


def _col(idx, keyword):
    """在(已归一化的)列名→索引字典中按包含关系找列（兼容『订单号/小票单号』合并表头）。"""
    nk = _norm(keyword)
    for k, v in idx.items():
        if nk in k:
            return v
    raise KeyError(keyword)


def _find_header(rows, *keywords):
    """返回 (index, header_row)：第一个每列(去空白后)能同时找到全部关键字的行。"""
    for i, r in enumerate(rows):
        cells = [_norm(c) for c in r if c is not None]
        if all(any(_norm(k) in c for c in cells) for k in keywords):
            return i, list(r)
    raise ValueError(f"找不到含 {keywords} 的表头行")


def _parse_date(s):
    if isinstance(s, datetime):
        return s.date()
    s = str(s)
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"无法解析日期: {s}")


def load_products_from_rows(info_rows, cost_rows=None):
    """商品信息表 + 销售成本表 → {barcode: Product}，按条码合并成本。"""
    cost_rows = cost_rows or []
    ii, ih = _find_header(info_rows, "国际条码")
    idx = {_norm(c): k for k, c in enumerate(ih) if c is not None}
    bc_i, name_i, spec_i, cat_i = idx["国际条码"], idx["商品名称"], idx["规格"], idx["类别"]
    cost_map = {}
    if cost_rows:
        ci, ch = _find_header(cost_rows, "商品条码")
        cidx = {_norm(c): k for k, c in enumerate(ch) if c is not None}
        cbc, ccost = cidx["商品条码"], cidx["销售成本"]
        for r in cost_rows[ci + 1:]:
            if r and len(r) > cbc and r[cbc] not in (None, ""):
                cost_map[str(r[cbc])] = _D(r[ccost])
    products = {}
    for r in info_rows[ii + 1:]:
        if not r or len(r) <= bc_i or r[bc_i] in (None, ""):
            continue
        bc = str(r[bc_i])
        products[bc] = Product(bc, str(r[name_i]), str(r[spec_i]), str(r[cat_i]),
                               cost_map.get(bc))
    return products


def load_stores_from_rows(rows):
    """『2026.6全部』风格 sheet → (stores, targets)。表头需含 类别/组别/名称/本月目标(可含 主管)。"""
    ii, h = _find_header(rows, "类别", "名称", "本月目标")
    idx = {_norm(c): k for k, c in enumerate(h) if c is not None}
    sup_i = idx.get("主管")
    stores, targets = {}, {}
    for r in rows[ii + 1:]:
        if not r or r[idx["名称"]] in (None, ""):
            continue
        cls = str(r[idx["类别"]]).strip()
        if cls not in ("A", "B", "C", "D"):
            continue
        name = str(r[idx["名称"]]).strip()
        supervisor = str(r[sup_i]).strip() if (sup_i is not None and sup_i < len(r)) else ""
        stores[name] = Store(name, str(r[idx["组别"]]).strip(), cls, supervisor)
        targets[name] = _D(r[idx["本月目标"]])
    return stores, targets


def load_sales_from_rows(rows):
    """销售流水 → [SalesLine]。"""
    ii, h = _find_header(rows, "小票单号", "销售金额")
    idx = {_norm(c): k for k, c in enumerate(h) if c is not None}
    g = lambda k: idx.get(k)
    lines = []
    for r in rows[ii + 1:]:
        if r[g("序号")] in (None, "") or r[g("销售时间")] in (None, ""):
            continue  # 跳过无序号或无销售时间的行（如合计行/异常行）
        src = r[g("源单号")]
        # 真实数据中『营业员名称』恒为空，当班人=收银员（对应『正确收银员』表）
        sp_i = g("收银员名称")
        lines.append(SalesLine(
            receipt=str(r[g("小票单号")]),
            src_order=(str(src) if src not in (None, "") else None),
            store=str(r[g("机构名称")]),
            sale_date=_parse_date(r[g("销售时间")]),
            barcode=str(r[g("国际条码")]),
            product_name=str(r[g("商品名称")]),
            qty=_D(r[g("数量")]),
            amount=_D(r[g("销售金额")]),
            unit_price=_D(r[g("销售单价")]),
            is_return=(str(r[g("销售方式")]) == "退货"),
            is_online=(str(r[g("订单渠道")]) == "线上"),
            salesperson=(str(r[sp_i]) if (sp_i is not None and r[sp_i] is not None) else ""),
        ))
    return lines


def load_gift_keys_from_rows(rows):
    """让利明细(整份=赠送清单) → {(订单号, 国际条码)} 集合。"""
    ii, h = _find_header(rows, "订单号", "国际条码")
    idx = {_norm(c): k for k, c in enumerate(h) if c is not None}
    o, b = _col(idx, "订单号"), _col(idx, "国际条码")  # 表头可能是『订单号/小票单号』合并
    keys = set()
    for r in rows[ii + 1:]:
        if r and len(r) > o and r[o] not in (None, ""):
            keys.add((str(r[o]), str(r[b])))
    return keys


def _xlsx_rows(path, sheet=None):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def load_products_xlsx(info_path, cost_path=None):
    return load_products_from_rows(_xlsx_rows(info_path),
                                   _xlsx_rows(cost_path) if cost_path else None)


def load_stores_xlsx(path, sheet):
    return load_stores_from_rows(_xlsx_rows(path, sheet))


def load_sales_xlsx(path):
    return load_sales_from_rows(_xlsx_rows(path))


def load_gift_keys_xlsx(path):
    return load_gift_keys_from_rows(_xlsx_rows(path))

"""导出逐笔提成台账（T7.1）。

每条 SalesRecord（被 compute 物化的 DetailRow）= 台账一行：
- DetailRow 字段：person/store/sale_date/barcode/product_name/tier/bucket/rate/amount/commission/tag
- SalesRecord 全字段：receipt/src_order/qty/unit_price/salesperson/cashier/is_return/is_online/
                      original_store/original_date/transfer_reason/extra
- 派生：是否调班 = SalesRecord.original_store IS NOT NULL（真值源在 SalesRecord，
        不依赖 DetailRow.is_transferred 占位列）

零重算：rows 由路由层直接从物化表 JOIN 查出传入，本模块只负责写盘。
"""
import json
from datetime import date as _date
from decimal import Decimal

from salary_engine.exporter_helpers import _bucket_display

# 列定义：(表头, 取值 key 或派生标记)
# 顺序即写入顺序；包含全字段审计（去向标签 + 提成 + 调班信息 + 源 extra）
_COLUMNS = [
    ("归属人",       "person"),
    ("归属店",       "store"),
    ("归属日",       "sale_date"),
    ("条码",         "barcode"),
    ("商品名称",     "product_name"),
    ("去向标签",     "tag"),
    ("商品档位",     "tier"),
    ("达成档",       "bucket"),
    ("提成比例",     "rate"),
    ("金额",         "amount"),
    ("提成金额",     "commission"),
    ("小票号",       "receipt"),
    ("源单号",       "src_order"),
    ("数量",         "qty"),
    ("单价",         "unit_price"),
    ("营业员",       "salesperson"),
    ("收银员",       "cashier"),
    ("是否退货",     "is_return"),
    ("是否线上",     "is_online"),
    ("是否调班",     "__transferred__"),    # 派生：original_store is not None
    ("原门店",       "original_store"),
    ("原日期",       "original_date"),
    ("调整原因",     "transfer_reason"),
    ("源字段",       "__extra__"),          # SalesRecord.extra JSON dump
]


_TIERS = ["常温高毛", "常温低毛", "低温高毛", "低温低毛", "特价"]


def build_summary_rows(results, tier_rows, store_map, target_map, duty_days_map, days):
    """从 Result 聚合 + 逐笔行构建 Sheet1「计算结果」行列表（26 列，与 CLI Sheet1 对齐）。
    排序：person 按总提成降序，store 按提成降序（与 CLI exporter_helpers 一致）。
    """
    from collections import defaultdict
    from decimal import Decimal

    # 5 档聚合：{(person, store, tier): [amount, commission, rate]}
    # 注：tier_rows 可能来自 text() SQL（金额为 float）或测试（Decimal），统一归一化。
    tier_agg = {}
    for r in tier_rows:
        tier = r.get("tier")
        if not tier or tier not in _TIERS:
            continue
        key = (r["person"], r["store"], tier)
        amount = Decimal(str(r.get("amount") or 0))
        commission = Decimal(str(r.get("commission") or 0))
        if key not in tier_agg:
            tier_agg[key] = [amount, commission, r.get("rate")]
        else:
            tier_agg[key][0] += amount
            tier_agg[key][1] += commission

    persons = defaultdict(list)
    for res in results:
        persons[res.person].append(res)
    person_total = {p: sum((r.commission or 0) for r in rs) for p, rs in persons.items()}

    out = []
    for person in sorted(person_total, key=lambda p: float(person_total[p]), reverse=True):
        stores_data = sorted(persons[person], key=lambda r: float(r.commission or 0), reverse=True)
        for res in stores_data:
            store = res.store
            store_class = store_map[store].store_class if store in store_map else ""
            monthly_target = Decimal(target_map.get(store, 0) or 0)
            daily_target = monthly_target / days if days else Decimal(0)
            duty = duty_days_map.get((person, store), 0)
            actual_target = daily_target * duty
            row = [
                person, store, store_class,
                round(float(monthly_target), 2), round(float(daily_target), 2),
                duty, round(float(actual_target), 2),
                round(float(res.sales or 0), 2),
                round(float(res.achievement or 0) * 100, 1),
                _bucket_display(res.bucket), round(float(res.commission or 0), 2),
            ]
            for tier in _TIERS:
                amount, comm, rate = tier_agg.get((person, store, tier), (Decimal(0), Decimal(0), None))
                rate_pct = f"{float(rate) * 100:.1f}%" if rate is not None else ""
                row.extend([round(float(amount), 2), rate_pct, round(float(comm), 2)])
            out.append(row)
    return out


def build_duty_grid(duty_rows):
    """考勤排版矩阵：{store: {day(int): salesperson}}，与前端 DutyGrid 一致。"""
    grid = {}
    for d in duty_rows:
        grid.setdefault(d.store, {})[d.duty_date.day] = d.salesperson
    return grid


def _fmt(v, key):
    """字段值渲染为 Excel 单元格值。None → ""；日期 → ISO；Decimal → float；
    派生字段（__transferred__/__extra__）按规则计算。"""
    if key == "__transferred__":
        return "是" if v else ""
    if key == "__extra__":
        if not v:
            return ""
        try:
            return json.dumps(v, ensure_ascii=False, default=str, sort_keys=True)
        except (TypeError, ValueError):
            return str(v)
    if v is None:
        return ""
    if isinstance(v, _date):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, bool):
        return "是" if v else ""
    return v


_SUMMARY_HEADERS = [
    "员工姓名", "门店", "门店类型", "月目标", "日目标", "考勤天数", "实际目标",
    "销售额", "达标率", "达标档位", "提成金额",
    "常温高毛_销售", "常温高毛_比例", "常温高毛_提成",
    "常温低毛_销售", "常温低毛_比例", "常温低毛_提成",
    "低温高毛_销售", "低温高毛_比例", "低温高毛_提成",
    "低温低毛_销售", "低温低毛_比例", "低温低毛_提成",
    "特价_销售", "特价_比例", "特价_提成",
]


def _style_header(ws, ncols):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, Side
    font = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    for cell in ws[1]:
        cell.font = font
        cell.alignment = align
        cell.border = thin
    for col in range(1, ncols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12


def _write_summary_sheet(ws, summary_rows):
    ws.append(_SUMMARY_HEADERS)
    _style_header(ws, len(_SUMMARY_HEADERS))
    for row in summary_rows:
        ws.append(row)
    ws.freeze_panes = "A2"


def _write_ledger_sheet(ws, rows):
    import openpyxl
    headers = [h for h, _ in _COLUMNS]
    ws.append(headers)
    _style_header(ws, len(_COLUMNS))
    for r in rows:
        # 派生：是否调班（original_store 非空 → 调班）
        transferred = bool(r.get("original_store"))
        row_vals = []
        for _, key in _COLUMNS:
            if key == "__transferred__":
                row_vals.append(_fmt(transferred, key))
            elif key == "__extra__":
                row_vals.append(_fmt(r.get("extra"), key))
            else:
                row_vals.append(_fmt(r.get(key), key))
        ws.append(row_vals)
    # 列宽：标签/审计列宽一点
    for idx, (hdr, _) in enumerate(_COLUMNS, start=1):
        if hdr in ("源字段", "调整原因", "商品名称"):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 32
        elif hdr in ("去向标签", "商品档位", "达成档", "原门店", "原日期"):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 14
    ws.freeze_panes = "A2"


def _write_duty_sheet(ws, duty_grid, days):
    ws.append(["门店"] + [f"{d}号" for d in range(1, days + 1)])
    for store in sorted(duty_grid):
        ws.append([store] + [duty_grid[store].get(d, "") for d in range(1, days + 1)])
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 14


def write_salary_export(month, summary_rows, ledger_rows, duty_grid, days, path):
    """写三 sheet 工资导出（ADR-020）：计算结果 + 提成台账 + 考勤排版。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "计算结果"
    _write_summary_sheet(ws1, summary_rows)
    ws2 = wb.create_sheet(f"提成台账-{month}")
    _write_ledger_sheet(ws2, ledger_rows)
    ws3 = wb.create_sheet("考勤排版")
    _write_duty_sheet(ws3, duty_grid, days)
    wb.save(path)

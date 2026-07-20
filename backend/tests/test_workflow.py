import openpyxl


def auth_header(client):
    t = client.post("/auth/login", json={"username": "admin", "password": "admin"}).json()["token"]
    return {"Authorization": f"Bearer {t}"}


def _sales_xlsx(path):
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["序号", "机构名称", "小票单号", "销售时间", "上传时间", "销售方式", "商品编码",
               "收银员名称", "国际条码", "数量", "销售金额", "销售单价", "商品名称",
               "订单渠道", "源单号"])
    ws.append(["1", "福景店", "R001", "2026-06-01 10:00", "", "销售", "", "高睿",
               "6920001", "1", "3", "3", "低温奶", "线下", ""])
    wb.save(path)


def _policy_content():
    """构造工资策略内容（commission_rates 存百分数，ADR-009）。"""
    from salary_engine.rates import _RATES, _TIERS
    cr = {}
    for cls, by_bucket in _RATES.items():
        for bucket, vals in by_bucket.items():
            for tier, val in zip(_TIERS, vals):
                cr.setdefault(cls, {}).setdefault(bucket, {})[tier] = str(val)
    return {"margin_rules": {}, "commission_rates": cr}


def _setup_computed_month(tmp_path, client, h):
    """建一个已计算完成的月份（建月→激活策略→门店/商品/目标→导入销售→排班→compute），
    返回 month 字符串。供多个 /compute 相关测试复用，消除三处脚手架重复。"""
    client.post("/months", headers=h, json={"month": "2026-06"})
    # 创建并激活工资策略（compute 读 SalaryPolicyVersion，策略存百分数；ADR-009）
    client.post("/salary-policies", headers=h, json={
        "effective_from": "2026-01-01", "content": _policy_content()})
    client.put("/stores/福景店", headers=h, json={"name": "福景店", "group": "1组", "store_class": "A"})
    client.put("/products/6920001", headers=h, json={
        "barcode": "6920001", "name": "低温奶", "spec": "200ml", "category": "低温奶", "cost": "2"})
    client.put("/months/2026-06/targets", headers=h, json={"items": [{"store": "福景店", "target": "3"}]})
    s = tmp_path / "sales.xlsx"; _sales_xlsx(s)
    with open(s, "rb") as f:
        client.post("/months/2026-06/import-sales", headers=h, files={"file": ("sales.xlsx", f)})
    client.post("/months/2026-06/infer-duty", headers=h)
    client.put("/months/2026-06/duty", headers=h, json={
        "items": [{"store": "福景店", "date": "2026-06-01", "salesperson": "高睿"}]})
    client.post("/months/2026-06/compute", headers=h)
    return "2026-06"


def test_import_sales_and_gifts(tmp_path, client):
    h = auth_header(client)
    client.post("/months", headers=h, json={"month": "2026-06"})
    s = tmp_path / "sales.xlsx"; _sales_xlsx(s)
    with open(s, "rb") as f:
        r = client.post("/months/2026-06/import-sales", headers=h,
                        files={"file": ("sales.xlsx", f)})
    assert r.status_code == 200
    m = client.get("/months/2026-06", headers=h).json()
    assert m["sales_file"] and m["sales_file"].endswith(".xlsx")


def test_infer_and_confirm_duty(tmp_path, client):
    h = auth_header(client)
    client.post("/months", headers=h, json={"month": "2026-06"})
    client.put("/stores/福景店", headers=h, json={"name": "福景店", "group": "1组", "store_class": "A"})
    s = tmp_path / "sales.xlsx"; _sales_xlsx(s)
    with open(s, "rb") as f:
        client.post("/months/2026-06/import-sales", headers=h, files={"file": ("sales.xlsx", f)})
    grid = client.post("/months/2026-06/infer-duty", headers=h).json()
    assert "福景店" in grid and "2026-06-01" in grid["福景店"]
    # 确认
    r = client.put("/months/2026-06/duty", headers=h, json={
        "items": [{"store": "福景店", "date": "2026-06-01", "salesperson": "高睿"}]})
    assert r.status_code == 200
    got = client.get("/months/2026-06/duty", headers=h).json()
    assert got["福景店"]["2026-06-01"] == "高睿"


def test_compute_and_result(tmp_path, client):
    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)
    res = client.get("/months/2026-06/results", headers=h).json()
    # 目标3/天0.1，卖3→达成3000%→GE_100；A低温高毛(单价3成本2=33%>15%)13%→0.39
    assert any(x["person"] == "高睿" and abs(x["commission"] - 0.39) < 0.01 for x in res["salary"])


def test_results_and_export(tmp_path, client):
    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)  # 复用：已算好 2026-06
    res = client.get("/months/2026-06/results", headers=h).json()
    assert res["salary"] and res["breakdown"]
    exp = client.get("/months/2026-06/export", headers=h)
    assert exp.status_code == 200
    ct = exp.headers.get("content-type", "")
    assert "spreadsheet" in ct or ct.startswith("application/vnd")


def test_results_stale_flag(tmp_path, client, db_session):
    """T4.3：/results 必须返回 stale 标志，让前端判断是否提示"数据已变更，请重新计算"。
    - 已计算 & 未变更 → stale=False
    - results_stale=True → stale=True（模拟 T5.1 的输入变更触发）
    - 未计算的草稿月份（status=draft, results_stale=True）→ stale=True
    """
    from backend.app.db import Month

    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)  # 建 2026-06 已计算月份

    # 已计算 & 未变更 → stale=False
    r = client.get("/months/2026-06/results", headers=h)
    assert r.status_code == 200, r.text
    assert r.json()["stale"] is False

    # 模拟 T5.1 的输入变更触发：直接翻 results_stale（写侧逻辑在 T5.1 落地）
    db_session.get(Month, "2026-06").results_stale = True
    db_session.commit()
    r2 = client.get("/months/2026-06/results", headers=h)
    assert r2.status_code == 200, r2.text
    assert r2.json()["stale"] is True

    # 未计算的草稿月份（默认 status=draft、results_stale=True）→ stale=True
    client.post("/months", headers=h, json={"month": "2026-07"})
    r3 = client.get("/months/2026-07/results", headers=h)
    assert r3.status_code == 200, r3.text
    assert r3.json()["stale"] is True


def test_compute_materializes_result_and_detail(tmp_path, client, db_session):
    """T4.1：/compute 必须同事务物化 Result（聚合）+ DetailRow（逐行），Σ 提成相等，
    status=computed、results_stale=False、policy_version_id 仅首次锁定。"""
    from decimal import Decimal
    from backend.app.db import Result, DetailRow, Month

    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)

    # Result 与 DetailRow 均落库
    assert db_session.query(Result).filter_by(month="2026-06").count() > 0
    assert db_session.query(DetailRow).filter_by(month="2026-06").count() > 0

    # Σ 不变量：逐行台账提成 == 聚合表提成
    dsum = sum((d.commission for d in db_session.query(DetailRow).filter_by(month="2026-06").all()),
               Decimal(0))
    rsum = sum((x.commission for x in db_session.query(Result).filter_by(month="2026-06").all()),
               Decimal(0))
    assert dsum == rsum, f"Σ mismatch: detail={dsum} result={rsum}"

    # 状态与首次锁定策略
    m = db_session.get(Month, "2026-06")
    assert m.status == "computed"
    assert m.results_stale is False
    assert m.policy_version_id is not None


def test_tier_summary_does_not_recompute(tmp_path, client, monkeypatch):
    """T4.2：tier_summary 必须读物化 DetailRow，零 _run_compute 调用（治 R1/R2 + C1）。"""
    from unittest.mock import MagicMock
    import backend.app.routers.workflow as workflow

    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)

    spy = MagicMock()
    monkeypatch.setattr(workflow, "_run_compute", spy)

    r = client.get("/months/2026-06/tier-summary", headers=h,
                   params={"store": "福景店", "person": "高睿"})
    assert r.status_code == 200, r.text
    assert spy.call_count == 0, "tier_summary 不应触发全量重算（应读物化 DetailRow）"
    data = r.json()
    assert isinstance(data.get("tiers"), list)
    # 与 compute 一致：A 类 GE_100 低温高毛 13% × 3 = 0.39
    assert abs(data["total_commission"] - 0.39) < 0.01, data
    # rate 已是 SalaryPolicyVersion 口径（13% 物化），不应再读 legacy RateVersion
    assert any(t["name"] == "低温高毛" and abs(t["rate"] - 0.13) < 0.001 for t in data["tiers"])


def test_tier_detail_does_not_recompute(tmp_path, client, monkeypatch):
    """T4.2：tier_detail 必须读物化 DetailRow，零 _run_compute 调用。"""
    from unittest.mock import MagicMock
    import backend.app.routers.workflow as workflow

    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)

    spy = MagicMock()
    monkeypatch.setattr(workflow, "_run_compute", spy)

    r = client.get("/months/2026-06/tier-detail", headers=h,
                   params={"store": "福景店", "person": "高睿", "bucket": "低温高毛"})
    assert r.status_code == 200, r.text
    assert spy.call_count == 0, "tier_detail 不应触发全量重算（应读物化 DetailRow）"
    items = r.json()["items"]
    assert items, "应当返回低温高毛档位的明细行"
    # 前端契约：tier_detail 必须返回完整 SalesItem（与 sales_detail 一致），而非 SLIM 子集
    it = items[0]
    expected_keys = {"id", "receipt", "src_order", "store", "sale_date", "barcode",
                     "product_name", "qty", "amount", "unit_price", "salesperson",
                     "cashier", "is_return", "is_online", "tag", "original_store",
                     "original_date", "transfer_reason"}
    assert expected_keys.issubset(it.keys()), f"缺字段: {expected_keys - set(it.keys())}"
    # 关键字段不再是 SLIM 形态独有：receipt/qty/unit_price 来自 JOIN 的 SalesRecord
    assert it["receipt"] == "R001"
    assert it["qty"] == 1.0
    assert it["unit_price"] == 3.0
    assert it["barcode"] == "6920001"


def test_input_changes_mark_month_stale(client, db_session, tmp_path):
    """T5.1：每个输入变更端点必须把 Month.results_stale 置 True（写侧），
    让 /results 的 stale 标志提示前端"数据已变更，请重新计算"。
    覆盖：set_duty / transfer_duty / set_targets / import_sales / import_gifts。
    salary_policies.activate 不在内 —— 已计算月份策略锁定，新策略不改其冻结结果（T4.1/H10）。"""
    from backend.app.db import Month

    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)  # 建 2026-06 已计算月份

    def _reset_stale():
        db_session.get(Month, "2026-06").results_stale = False
        db_session.commit()

    def _assert_stale(label):
        db_session.expire_all()
        m = db_session.get(Month, "2026-06")
        assert m.results_stale is True, f"{label} 后 results_stale 应为 True"

    # set_duty (PUT /months/{m}/duty)
    _reset_stale()
    r = client.put("/months/2026-06/duty", headers=h, json={
        "items": [{"store": "福景店", "date": "2026-06-01", "salesperson": "高睿"}]})
    assert r.status_code == 200, r.text
    _assert_stale("set_duty")

    # transfer_duty (POST /months/{m}/duty/transfer)
    _reset_stale()
    r = client.post("/months/2026-06/duty/transfer", headers=h, json={
        "from_store": "福景店", "to_store": "新店",
        "date": "2026-06-01", "salesperson": "高睿"})
    assert r.status_code == 200, r.text
    _assert_stale("transfer_duty")

    # set_targets (PUT /months/{m}/targets)
    _reset_stale()
    r = client.put("/months/2026-06/targets", headers=h, json={
        "items": [{"store": "福景店", "target": "5"}]})
    assert r.status_code == 200, r.text
    _assert_stale("set_targets")

    # import_sales (POST /months/{m}/import-sales)
    _reset_stale()
    s = tmp_path / "sales2.xlsx"; _sales_xlsx(s)
    with open(s, "rb") as f:
        r = client.post("/months/2026-06/import-sales", headers=h,
                        files={"file": ("sales.xlsx", f)})
    assert r.status_code == 200, r.text
    _assert_stale("import_sales")

    # import_gifts (POST /months/{m}/import-gifts)
    _reset_stale()
    g = tmp_path / "gifts.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["订单号", "国际条码"])
    ws.append(["R001", "6920001"])
    wb.save(g)
    with open(g, "rb") as f:
        r = client.post("/months/2026-06/import-gifts", headers=h,
                        files={"file": ("gifts.xlsx", f)})
    assert r.status_code == 200, r.text
    _assert_stale("import_gifts")


def test_compute_single_flight(tmp_path, client, db_session):
    """T5.2（治 R3）：do_compute 必须按月加锁；同一月份已有计算进行中时，
    第二次 /compute 应立即返回 409，而不是叠加 CPU-bound 工作。
    模拟方式：测试线程直接持有该月的锁，再调 /compute 应得 409。"""
    import backend.app.routers.workflow as wf

    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)   # 确保月份存在且已计算

    # 模拟另一个正在进行的 compute：测试线程持有该月的锁
    lock = wf._get_lock("2026-06")
    # setdefault 保证同月份两次取锁是同一实例（消除 check-then-act 竞态）
    assert wf._get_lock("2026-06") is lock
    got = lock.acquire(blocking=False)
    assert got
    try:
        r = client.post("/months/2026-06/compute", headers=h)
        assert r.status_code == 409, f"并发计算应返回 409，实际 {r.status_code}: {r.text}"
    finally:
        lock.release()


def test_compute_restores_status_on_failure(tmp_path, client, db_session, monkeypatch):
    """T5.2：计算失败时必须把 status 从 "computing" 恢复到先前值，避免月份卡死。"""
    import pytest
    from backend.app.db import Month
    import backend.app.routers.workflow as wf

    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)

    # 先记下先前状态（应为 computed）
    db_session.expire_all()
    prev = db_session.get(Month, "2026-06").status
    assert prev == "computed"

    # 让 _run_compute 抛异常，模拟计算失败
    def _boom(db, month):
        raise RuntimeError("engine exploded")
    monkeypatch.setattr(wf, "_run_compute", _boom)

    # Starlette TestClient 默认会把服务端异常原样抛出
    with pytest.raises(RuntimeError):
        client.post("/months/2026-06/compute", headers=h)

    # 状态应恢复成 computed，而非卡在 computing
    db_session.expire_all()
    m = db_session.get(Month, "2026-06")
    assert m.status == prev, f"失败后 status 应恢复为 {prev}，实际 {m.status}"
    assert m.status != "computing", "失败后不应卡在 computing"


def test_export_reads_materialized_no_recompute(tmp_path, client, monkeypatch):
    """T7.1：/export 必须读物化 DetailRow JOIN SalesRecord，零 _run_compute 调用（治 R1）。
    每条导入记录 = 台账一行；去向标签 + 提成金额从物化表直接读取。"""
    from unittest.mock import MagicMock
    import backend.app.routers.workflow as workflow

    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)

    spy = MagicMock()
    monkeypatch.setattr(workflow, "_run_compute", spy)

    r = client.get("/months/2026-06/export", headers=h)
    assert r.status_code == 200, r.text
    assert spy.call_count == 0, "export 不应触发全量重算（应读物化 DetailRow JOIN SalesRecord）"
    ct = r.headers.get("content-type", "")
    assert "spreadsheet" in ct or ct.startswith("application/vnd"), f"bad ct: {ct}"
    assert len(r.content) > 0, "导出体不应为空"


def test_export_ledger_has_tags_and_fields(tmp_path, client):
    """T7.1：导出台账每条记录一行，含去向标签 + 提成金额 + 全字段（含源 extra）。
    解析 xlsx，校验表头与至少一行数据带已知 fate 标签。"""
    import io
    import openpyxl

    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)

    r = client.get("/months/2026-06/export", headers=h)
    assert r.status_code == 200, r.text

    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    assert "提成台账" in wb.sheetnames, f"sheet names: {wb.sheetnames}"
    ws = wb["提成台账"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) >= 2, f"台账至少表头 + 1 行数据，实际 {len(rows)}"
    headers = list(rows[0])
    header_set = set(headers)
    # 关键列必须存在
    for required in ("去向标签", "提成金额", "商品档位", "达成档", "提成比例",
                     "是否调班", "源字段"):
        assert required in header_set, f"缺少列 {required}: {headers}"

    tag_idx = headers.index("去向标签")
    comm_idx = headers.index("提成金额")
    data_rows = rows[1:]
    tags = {row[tag_idx] for row in data_rows if row[tag_idx]}
    known_tags = {"有效计提", "退货冲抵", "退货未匹配", "赠送剔除", "不计提成", "非乳品"}
    assert tags & known_tags, f"无已知 fate 标签: {tags}"
    # 至少一行有非空提成金额
    assert any(row[comm_idx] not in (None, 0, 0.0) for row in data_rows), "无提成金额"


def test_compute_recompute_preserves_policy_lock(tmp_path, client, db_session):
    """H10：首次 /compute 锁定 policy_version_id；策略变更后再次 /compute 不应覆盖锁定。
    若 do_compute 里 `if m.policy_version_id is None` 守卫被删除，本测试必须失败。"""
    from backend.app.db import Month

    h = auth_header(client)
    _setup_computed_month(tmp_path, client, h)

    # 首次 compute 后，月份已锁定到 v1
    db_session.expire_all()
    m = db_session.get(Month, "2026-06")
    v1_id = m.policy_version_id
    assert v1_id is not None, "首次 compute 应锁定 policy_version_id"

    # 新建第二个工资策略（/salary-policies 会把 is_current 切到新版本，即 v2）
    r2 = client.post("/salary-policies", headers=h, json={
        "effective_from": "2026-07-01", "content": _policy_content()})
    assert r2.status_code == 201
    v2_id = r2.json()["id"]
    assert v2_id != v1_id

    # 再次 /compute：守卫 `if m.policy_version_id is None` 应使其跳过覆盖，仍保持 v1
    r = client.post("/months/2026-06/compute", headers=h)
    assert r.status_code == 200

    db_session.expire_all()
    m = db_session.get(Month, "2026-06")
    assert m.policy_version_id == v1_id, (
        f"重算不应覆盖已锁定的 policy_version_id: expected v1={v1_id}, got {m.policy_version_id}")
    assert m.policy_version_id != v2_id

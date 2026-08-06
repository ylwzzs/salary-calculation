from datetime import date
from decimal import Decimal
from types import SimpleNamespace

from backend.app.services.ledger_export import build_summary_rows


def _result(person, store, sales, target, achievement, bucket, commission):
    return SimpleNamespace(
        person=person, store=store,
        sales=Decimal(sales), target=Decimal(target),
        achievement=Decimal(achievement), bucket=bucket,
        commission=Decimal(commission),
    )


def test_build_summary_rows_single_row():
    results = [_result("高睿", "福景店", "10", "1", "1", "GE_100", "1.4")]
    tier_rows = [
        {"person": "高睿", "store": "福景店", "tier": "低温高毛",
         "rate": Decimal("0.14"), "amount": Decimal("10"), "commission": Decimal("1.4")},
    ]
    store_map = {"福景店": SimpleNamespace(store_class="A")}
    target_map = {"福景店": Decimal("30")}
    duty_days_map = {("高睿", "福景店"): 1}

    rows = build_summary_rows(results, tier_rows, store_map, target_map, duty_days_map, 30)

    assert len(rows) == 1
    row = rows[0]
    # 前 11 列：人/店/类型/月目标/日目标/考勤/实际目标/销售/达标率/档位/提成
    assert row[:11] == [
        "高睿", "福景店", "A", 30.0, 1.0, 1, 1.0, 10.0, 100.0, "≥100%", 1.4,
    ]
    # 低温高毛 idx 17/18/19：销售 10、比例 14.0%、提成 1.4
    assert row[17] == 10.0
    assert row[18] == "14.0%"
    assert row[19] == 1.4
    # 其它档位为 0 / 空
    assert row[11] == 0.0 and row[13] == 0.0 and row[23] == 0.0


def test_write_salary_export_totals_row(tmp_path):
    """Sheet1 底部合计行：销售额/提成金额求和 + 数值 2 位小数。"""
    import openpyxl
    from backend.app.services.ledger_export import write_salary_export, build_summary_rows

    results = [
        _result("甲", "店1", "10", "1", "1", "GE_100", "1.4"),
        _result("乙", "店1", "20", "1", "1", "GE_100", "2.6"),
    ]
    summary = build_summary_rows(results, [], {}, {}, {}, 30)
    out = tmp_path / "totals.xlsx"
    write_salary_export("2026-06", summary, [], {}, 30, str(out))

    wb = openpyxl.load_workbook(out)
    ws = wb["计算结果"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    last = rows[-1]
    assert last[0] == "合计", f"末行应为合计: {last[0]}"
    sales_i = headers.index("销售额")
    comm_i = headers.index("提成金额")
    assert abs((last[sales_i] or 0) - 30.0) < 0.01, last[sales_i]
    assert abs((last[comm_i] or 0) - 4.0) < 0.01, last[comm_i]


def test_build_summary_rows_orders_by_commission_desc():
    results = [
        _result("甲", "店1", "1", "1", "1", "GE_100", "2"),
        _result("乙", "店1", "1", "1", "1", "GE_100", "5"),
    ]
    rows = build_summary_rows(results, [], {}, {}, {}, 30)
    assert [r[0] for r in rows] == ["乙", "甲"]


def test_build_duty_grid():
    from backend.app.services.ledger_export import build_duty_grid
    d1 = SimpleNamespace(store="福景店", duty_date=date(2026, 6, 1), salesperson="高睿")
    d2 = SimpleNamespace(store="福景店", duty_date=date(2026, 6, 2), salesperson="张三")
    d3 = SimpleNamespace(store="螺农店", duty_date=date(2026, 6, 1), salesperson="李四")
    assert build_duty_grid([d1, d2, d3]) == {
        "福景店": {1: "高睿", 2: "张三"},
        "螺农店": {1: "李四"},
    }


def test_write_salary_export_three_sheets(tmp_path):
    import openpyxl
    from backend.app.services.ledger_export import (
        write_salary_export, build_summary_rows, build_duty_grid)

    results = [_result("高睿", "福景店", "10", "1", "1", "GE_100", "1.4")]
    tier_rows = [{"person": "高睿", "store": "福景店", "tier": "低温高毛",
                  "rate": Decimal("0.14"), "amount": Decimal("10"), "commission": Decimal("1.4")}]
    summary = build_summary_rows(
        results, tier_rows,
        {"福景店": SimpleNamespace(store_class="A")},
        {"福景店": Decimal("30")}, {("高睿", "福景店"): 1}, 30)
    ledger = [{
        "person": "高睿", "store": "福景店", "sale_date": date(2026, 6, 1),
        "barcode": "6920001", "product_name": "低温奶", "tag": "有效计提",
        "tier": "低温高毛", "bucket": "GE_100", "rate": Decimal("0.14"),
        "amount": Decimal("10"), "commission": Decimal("1.4"),
        "receipt": "R001", "src_order": None, "qty": 1, "unit_price": Decimal("10"),
        "salesperson": "高睿", "cashier": "", "is_return": False, "is_online": False,
        "original_store": None, "original_date": None, "transfer_reason": None, "extra": None,
    }]
    duty_rows = [SimpleNamespace(store="福景店", duty_date=date(2026, 6, 1), salesperson="高睿")]
    duty_grid = build_duty_grid(duty_rows)
    out = tmp_path / "out.xlsx"
    write_salary_export("2026-06", summary, ledger, duty_grid, 30, str(out))

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["计算结果", "提成台账-2026-06", "考勤排版"], wb.sheetnames

    ws1 = wb["计算结果"]
    h1 = list(ws1.iter_rows(values_only=True))
    assert h1[0][0] == "员工姓名" and h1[0][10] == "提成金额"
    assert h1[1][0] == "高睿" and h1[1][10] == 1.4

    ws2 = wb["提成台账-2026-06"]
    h2 = list(ws2.iter_rows(values_only=True))
    assert h2[0][0] == "归属人" and h2[0][2] == "归属日"
    assert h2[1][0] == "高睿" and h2[1][1] == "福景店"

    ws3 = wb["考勤排版"]
    h3 = list(ws3.iter_rows(values_only=True))
    assert h3[0][0] == "门店" and h3[0][1] == "1号" and h3[0][30] == "30号"
    assert h3[1][0] == "福景店" and h3[1][1] == "高睿"

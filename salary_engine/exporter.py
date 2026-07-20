"""导出工资表与明细（结果页两个sheet）。"""
from collections import defaultdict
from decimal import Decimal
from salary_engine.calculator import ComputeResult


def write_excel(result: ComputeResult, out_path: str, month: str = None, db=None):
    """用 openpyxl 写两个 sheet：
    Sheet1: 计算结果列表 + 档位提成明细列
    Sheet2: 全部销售明细

    H12：所有 backend.* 导入延迟到 `if db:` 块内，使 db=None（CLI 独立运行）
    时完全不触碰 backend 包，保持 salary_engine 自包含可独立使用。
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, Font
    from salary_engine.margin import gross_margin, classify_tier

    wb = openpyxl.Workbook()

    # ============ Sheet1: 计算结果 ============
    ws1 = wb.active
    ws1.title = "计算结果"

    # 表头
    headers1 = [
        "员工姓名", "门店", "门店类型", "月目标", "日目标",
        "考勤天数", "实际目标", "销售额", "达标率", "达标档位", "提成金额",
        # 档位明细列
        "常温高毛_销售", "常温高毛_比例", "常温高毛_提成",
        "常温低毛_销售", "常温低毛_比例", "常温低毛_提成",
        "低温高毛_销售", "低温高毛_比例", "低温高毛_提成",
        "低温低毛_销售", "低温低毛_比例", "低温低毛_提成",
        "特价_销售", "特价_比例", "特价_提成",
    ]
    ws1.append(headers1)

    # 表头样式
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in ws1[1]:
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # 从数据库获取门店信息、费率表、考勤天数、月度目标
    store_map = {}      # name -> Store
    rate_rates = {}     # rate_table.rates: {(cls, bucket, tier): Decimal}
    duty_days_map = {}  # (person, store) -> 天数
    target_map = {}     # store -> Decimal(月目标)

    if db:
        # H12：backend.* 导入全部在 db 分支内（CLI db=None 时不触碰 backend）
        from backend.app.db import (
            RateVersion, Month as MonthModel,
            SalesRecord, Duty, Product, Store, MonthlyTarget,
        )

        for s in db.query(Store).all():
            store_map[s.name] = s

        # 加载费率表（与计算时锁定版本一致）
        m = db.get(MonthModel, month)
        if m and m.rate_version_id:
            rv = db.get(RateVersion, m.rate_version_id)
        else:
            rv = db.query(RateVersion).filter_by(is_current=True).first()
        if rv and rv.rates:
            for cls, by_bucket in rv.rates.items():
                for bucket, by_tier in by_bucket.items():
                    for tier, pct in by_tier.items():
                        rate_rates[(cls, bucket, tier)] = Decimal(str(pct))

        # 统计每个 (person, store) 的出勤天数
        for d in db.query(Duty).filter_by(month=month).all():
            key = (d.salesperson, d.store)
            duty_days_map[key] = duty_days_map.get(key, 0) + 1

        # 月度目标
        for t in db.query(MonthlyTarget).filter_by(month=month).all():
            target_map[t.store] = Decimal(t.target)

        # H12：engine_bridge 也是 backend 模块，仅 db 时导入；CLI 走 30 天默认
        from backend.app.services.engine_bridge import days_in_month as _dim
        total_days = _dim(month) if month else 30
    else:
        total_days = 30

    tier_names = ["常温高毛", "常温低毛", "低温高毛", "低温低毛", "特价"]

    # 构建行数据：按 person 分组，每个 person 下按 store 排序
    # result.breakdown: {(person, store): {sales, target, achievement, bucket, commission}}
    from salary_engine.exporter_helpers import build_rows_from_breakdown
    rows_data = build_rows_from_breakdown(
        result, store_map, rate_rates, duty_days_map, target_map, total_days, tier_names
    )

    # 写入行并合并同员工单元格
    row_idx = 2  # Excel 行号（第1行是表头）
    for person, stores_data in rows_data:
        start_row = row_idx
        for sd in stores_data:
            ws1.append(sd)
            # 设置单元格样式
            for col in range(1, len(headers1) + 1):
                cell = ws1.cell(row=row_idx, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_idx += 1

        # 合并同员工的姓名列（A列）
        end_row = row_idx - 1
        if end_row > start_row:
            ws1.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws1.cell(row=start_row, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    # 设置列宽
    col_widths = {
        1: 10, 2: 14, 3: 8, 4: 10, 5: 10,
        6: 8, 7: 10, 8: 12, 9: 8, 10: 8, 11: 12,
    }
    for col, w in col_widths.items():
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    for col in range(12, len(headers1) + 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

    # ============ Sheet2: 全部明细 ============
    ws2 = wb.create_sheet("销售明细")

    headers2 = [
        "标签", "小票号", "源单号", "门店", "日期", "条码", "商品名称",
        "单价", "数量", "金额", "营业员", "收银员", "退货", "线上",
        "原始门店", "原始日期", "调整原因",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    if month and db:
        from backend.app.db import SalesRecord
        records = db.query(SalesRecord).filter(SalesRecord.month == month).order_by(
            SalesRecord.store, SalesRecord.sale_date, SalesRecord.receipt
        ).all()

        for r in records:
            ws2.append([
                r.tag, r.receipt, r.src_order or "", r.store,
                r.sale_date.isoformat() if r.sale_date else "",
                r.barcode, r.product_name,
                round(float(r.unit_price), 2) if r.unit_price else 0,
                float(r.qty) if r.qty else 0,
                round(float(r.amount), 2) if r.amount else 0,
                r.salesperson, r.cashier or "",
                "是" if r.is_return else "", "是" if r.is_online else "",
                r.original_store or "",
                r.original_date.isoformat() if r.original_date else "",
                r.transfer_reason or "",
            ])

    # 设置列宽
    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    wb.save(out_path)

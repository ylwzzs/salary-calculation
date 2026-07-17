from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
from PIL import Image as PILImage

from backend.app.auth import current_user
from backend.app.db import get_db, SalaryPolicyVersion, Month, User
from backend.app.schemas import (
    SalaryPolicyOut,
    SalaryPolicyCreate,
    SalaryPolicySummary,
)

router = APIRouter(prefix="/salary-policies", tags=["salary-policies"])


def _month_usage(db: Session, policy_id: int) -> list[str]:
    return [
        row[0]
        for row in db.query(Month.month)
        .filter(Month.policy_version_id == policy_id)
        .order_by(Month.month)
        .all()
    ]


@router.get("", response_model=list[SalaryPolicySummary])
def list_policies(_: User = Depends(current_user), db: Session = Depends(get_db)):
    versions = (
        db.query(SalaryPolicyVersion)
        .order_by(SalaryPolicyVersion.version.desc())
        .all()
    )
    result = []
    for v in versions:
        summary = SalaryPolicySummary.model_validate(v)
        summary.used_by_months = _month_usage(db, v.id)
        result.append(summary)
    return result


@router.get("/current", response_model=SalaryPolicyOut)
def get_current(_: User = Depends(current_user), db: Session = Depends(get_db)):
    policy = db.query(SalaryPolicyVersion).filter_by(is_current=True).first()
    if policy is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "当前没有激活的薪酬制度版本")
    return policy


@router.get("/{policy_id}", response_model=SalaryPolicyOut)
def get_policy(policy_id: int, _: User = Depends(current_user), db: Session = Depends(get_db)):
    policy = db.get(SalaryPolicyVersion, policy_id)
    if policy is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "薪酬制度版本不存在")
    return policy


@router.post("", response_model=SalaryPolicyOut, status_code=status.HTTP_201_CREATED)
def create_policy(
    body: SalaryPolicyCreate,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    next_version = (db.query(func.max(SalaryPolicyVersion.version)).scalar() or 0) + 1

    for v in db.query(SalaryPolicyVersion).filter_by(is_current=True).all():
        v.is_current = False

    policy = SalaryPolicyVersion(
        version=next_version,
        effective_from=body.effective_from,
        is_current=True,
        created_by=user.username,
        content=body.content.model_dump(),
        note=body.note,
    )
    db.add(policy)
    db.commit()
    db.refresh(policy)
    return policy


@router.post("/{policy_id}/activate", response_model=SalaryPolicyOut)
def activate_policy(
    policy_id: int, _: User = Depends(current_user), db: Session = Depends(get_db)
):
    policy = db.get(SalaryPolicyVersion, policy_id)
    if policy is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "薪酬制度版本不存在")

    for v in db.query(SalaryPolicyVersion).filter_by(is_current=True).all():
        v.is_current = False

    policy.is_current = True
    db.commit()
    db.refresh(policy)
    return policy


@router.delete("/{policy_id}")
def delete_policy(
    policy_id: int, _: User = Depends(current_user), db: Session = Depends(get_db)
):
    policy = db.get(SalaryPolicyVersion, policy_id)
    if policy is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "薪酬制度版本不存在")

    if policy.is_current:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "不能删除当前激活的薪酬制度版本"
        )

    used_months = _month_usage(db, policy_id)
    if used_months:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"不能删除已被月份使用的薪酬制度版本: {', '.join(used_months)}",
        )

    total = db.query(SalaryPolicyVersion).count()
    if total <= 1:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "不能删除唯一剩余的薪酬制度版本"
        )

    db.delete(policy)
    db.commit()
    return {"ok": True}


@router.get("/{policy_id}/export/excel")
def export_excel(
    policy_id: int, _: User = Depends(current_user), db: Session = Depends(get_db)
):
    """导出薪酬制度为 Excel 文件"""
    policy = db.get(SalaryPolicyVersion, policy_id)
    if policy is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "薪酬制度版本不存在")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"薪酬制度v{policy.version}"

    # 样式定义
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # 标题行
    ws["A1"] = f"薪酬制度 v{policy.version}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A2"] = f"生效日期: {policy.effective_from}"
    ws["A3"] = f"备注: {policy.note or '无'}"

    # 毛利率规则表
    ws["A5"] = "一、毛利率分类规则"
    ws["A5"].font = header_font

    row = 6
    ws.cell(row=row, column=1, value="商品分类").font = header_font
    ws.cell(row=row, column=2, value="正价(高毛利)").font = header_font
    ws.cell(row=row, column=3, value="低价(低毛利)").font = header_font
    ws.cell(row=row, column=4, value="特价").font = header_font
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).border = thin_border

    row += 1
    margin_rules = policy.content.get("margin_rules", {})
    for cat, rules in margin_rules.items():
        high_min = rules.get("high", {}).get("min", "")
        low_min = rules.get("low", {}).get("min", "")
        low_max = rules.get("low", {}).get("max", "")
        special_max = rules.get("special", {}).get("max", "")

        ws.cell(row=row, column=1, value=cat).border = thin_border
        ws.cell(row=row, column=2, value=f">{high_min}%" if high_min else "-").border = thin_border
        ws.cell(row=row, column=3, value=f"{low_min}-{low_max}%" if low_min and low_max else "-").border = thin_border
        ws.cell(row=row, column=4, value=f"≤{special_max}%" if special_max else "-").border = thin_border
        row += 1

    # 提成比例表
    row += 2
    ws.cell(row=row, column=1, value="二、提成比例表(%)").font = header_font
    row += 1

    # 表头: 达成档位 | 商品档位 | A类 | B类 | C类 | D类
    ws.cell(row=row, column=1, value="达成档位").font = header_font
    ws.cell(row=row, column=2, value="商品档位").font = header_font
    ws.cell(row=row, column=3, value="A类").font = header_font
    ws.cell(row=row, column=4, value="B类").font = header_font
    ws.cell(row=row, column=5, value="C类").font = header_font
    ws.cell(row=row, column=6, value="D类").font = header_font
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).border = thin_border

    commission_rates = policy.content.get("commission_rates", {})
    buckets = ["GE_100", "90_100", "80_90", "70_80", "LT_70"]
    bucket_labels = [">=100%", "90-100%", "80-90%", "70-80%", "<70%"]
    tiers = ["低温低毛", "低温高毛", "常温低毛", "常温高毛", "特价"]

    row += 1
    for bi, bucket in enumerate(buckets):
        for ti, tier in enumerate(tiers):
            ws.cell(row=row, column=1, value=bucket_labels[bi] if ti == 0 else "").border = thin_border
            if ti == 0:
                ws.cell(row=row, column=1).alignment = Alignment(vertical="center")

            ws.cell(row=row, column=2, value=tier).border = thin_border

            for ci, cls in enumerate(["A", "B", "C", "D"]):
                val = commission_rates.get(cls, {}).get(bucket, {}).get(tier, "-")
                ws.cell(row=row, column=3 + ci, value=val).border = thin_border
                ws.cell(row=row, column=3 + ci).alignment = center_align

            row += 1

    # 调整列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 10

    # 输出
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = f"salary_policy_v{policy.version}_{policy.effective_from}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={quote(filename)}"},
    )


def _build_policy_elements(policy):
    """构建 PDF 元素列表（毛利率规则 + 提成比例表）"""
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("PolicyTitle", parent=styles["Title"], fontSize=16, spaceAfter=6)
    heading_style = ParagraphStyle("PolicyHeading", parent=styles["Heading2"], fontSize=12, spaceAfter=6, spaceBefore=10)
    meta_style = ParagraphStyle("PolicyMeta", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    elements = []
    elements.append(Paragraph(f"薪酬制度 v{policy.version}", title_style))
    elements.append(Paragraph(f"生效日期: {policy.effective_from}　　备注: {policy.note or '无'}", meta_style))
    elements.append(Spacer(1, 8))

    # 毛利率规则表
    margin_rules = policy.content.get("margin_rules", {})
    elements.append(Paragraph("一、毛利率分类规则", heading_style))

    margin_data = [["商品分类", "正价(高毛利)", "低价(低毛利)", "特价"]]
    for cat, rules in margin_rules.items():
        high_min = rules.get("high", {}).get("min", "")
        low_min = rules.get("low", {}).get("min", "")
        low_max = rules.get("low", {}).get("max", "")
        special_max = rules.get("special", {}).get("max", "")
        margin_data.append([
            cat,
            f">{high_min}%" if high_min else "-",
            f"{low_min}-{low_max}%" if low_min and low_max else "-",
            f"≤{special_max}%" if special_max else "-",
        ])

    margin_table = RLTable(margin_data, colWidths=[60, 100, 100, 80])
    margin_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT", (0, 0), (-1, -1), 22),
    ]))
    elements.append(margin_table)
    elements.append(Spacer(1, 12))

    # 提成比例表
    commission_rates = policy.content.get("commission_rates", {})
    elements.append(Paragraph("二、提成比例表(%)", heading_style))

    buckets = ["GE_100", "90_100", "80_90", "70_80", "LT_70"]
    bucket_labels = [">=100%", "90-100%", "80-90%", "70-80%", "<70%"]
    tiers = ["低温低毛", "低温高毛", "常温低毛", "常温高毛", "特价"]
    classes = ["A", "B", "C", "D"]

    rate_data = [["达成档位", "商品档位"] + [f"{c}类" for c in classes]]
    for bi, bucket in enumerate(buckets):
        for ti, tier in enumerate(tiers):
            row = []
            if ti == 0:
                row.append(bucket_labels[bi])
            else:
                row.append("")
            row.append(tier)
            for cls in classes:
                val = commission_rates.get(cls, {}).get(bucket, {}).get(tier, "-")
                row.append(str(val))
            rate_data.append(row)

    rate_table = RLTable(rate_data, colWidths=[65, 65, 45, 45, 45, 45])
    rate_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT", (0, 0), (-1, -1), 18),
        ("SPAN", (0, 1), (0, 5)),
        ("SPAN", (0, 6), (0, 10)),
        ("SPAN", (0, 11), (0, 15)),
        ("SPAN", (0, 16), (0, 20)),
        ("SPAN", (0, 21), (0, 25)),
    ]))
    elements.append(rate_table)

    return elements


@router.get("/{policy_id}/export/pdf")
def export_pdf(
    policy_id: int, _: User = Depends(current_user), db: Session = Depends(get_db)
):
    """导出薪酬制度为 PDF 文件"""
    policy = db.get(SalaryPolicyVersion, policy_id)
    if policy is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "薪酬制度版本不存在")

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm, topMargin=20 * mm, bottomMargin=15 * mm)
    elements = _build_policy_elements(policy)
    doc.build(elements)
    buf.seek(0)

    from urllib.parse import quote
    filename = f"salary_policy_v{policy.version}_{policy.effective_from}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={quote(filename)}"},
    )


@router.get("/{policy_id}/export/image")
def export_image(
    policy_id: int, _: User = Depends(current_user), db: Session = Depends(get_db)
):
    """导出薪酬制度为 PNG 图片"""
    policy = db.get(SalaryPolicyVersion, policy_id)
    if policy is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "薪酬制度版本不存在")

    # 先生成 PDF，再转成图片
    pdf_buf = BytesIO()
    doc = SimpleDocTemplate(pdf_buf, pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm, topMargin=20 * mm, bottomMargin=15 * mm)
    elements = _build_policy_elements(policy)
    doc.build(elements)
    pdf_buf.seek(0)

    # PDF 转 PNG（使用 fitz/PyMuPDF 或回退到 reportlab 渲染）
    try:
        import fitz  # PyMuPDF
        pdf_doc = fitz.open(stream=pdf_buf.read(), filetype="pdf")
        page = pdf_doc[0]
        mat = fitz.Matrix(2, 2)  # 2x 放大以提高清晰度
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        pdf_doc.close()
    except ImportError:
        # 无 PyMuPDF，使用 Pillow 从 PDF 渲染
        # 尝试用 pdf2image
        try:
            from pdf2image import convert_from_bytes
            images = convert_from_bytes(pdf_buf.read(), dpi=200)
            img_buf = BytesIO()
            images[0].save(img_buf, format="PNG")
            img_bytes = img_buf.getvalue()
        except ImportError:
            raise HTTPException(
                status.HTTP_501_NOT_IMPLEMENTED,
                "图片导出需要安装 PyMuPDF (pip install PyMuPDF) 或 pdf2image",
            )

    from urllib.parse import quote
    filename = f"salary_policy_v{policy.version}_{policy.effective_from}.png"
    return StreamingResponse(
        BytesIO(img_bytes),
        media_type="image/png",
        headers={"Content-Disposition": f"attachment; filename={quote(filename)}"},
    )

